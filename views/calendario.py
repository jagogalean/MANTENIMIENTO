import streamlit as st
import base64
import calendar
import pandas as pd
from io import BytesIO
from datetime import datetime, date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
DIAS_SEMANA_ABR = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]


def _dias_programados_en_mes(ultima_ejecucion_str, frecuencia_dias, anio, mes):
    """Proyección: qué días del mes le tocarían a la tarea según su frecuencia."""
    primer_dia_mes = date(anio, mes, 1)
    ultimo_dia_num = calendar.monthrange(anio, mes)[1]
    fin_mes = date(anio, mes, ultimo_dia_num)

    if not frecuencia_dias or frecuencia_dias <= 0:
        return set()

    if ultima_ejecucion_str:
        try:
            ancla = datetime.strptime(ultima_ejecucion_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            ancla = primer_dia_mes
    else:
        ancla = primer_dia_mes

    if ancla < primer_dia_mes:
        pasos = (primer_dia_mes - ancla).days // frecuencia_dias
        cursor = ancla + timedelta(days=pasos * frecuencia_dias)
        while cursor < primer_dia_mes:
            cursor += timedelta(days=frecuencia_dias)
    else:
        cursor = ancla

    dias = set()
    while cursor <= fin_mes:
        if cursor >= primer_dia_mes:
            dias.add(cursor.day)
        cursor += timedelta(days=frecuencia_dias)
    return dias


def _dias_realizados_en_mes(ejecuciones_plan, anio, mes):
    """Días del mes en que la tarea SE HIZO REALMENTE, según el historial
    (puede no coincidir con el día que estaba programado)."""
    dias = set()
    for e in ejecuciones_plan:
        fr = e.get("fecha_realizada")
        if not fr:
            continue
        try:
            f = datetime.strptime(fr, "%Y-%m-%d").date()
        except ValueError:
            continue
        if f.year == anio and f.month == mes:
            dias.add(f.day)
    return dias


def _preparar_tareas(planes, ejecuciones, maquinas_filtradas_ids=None):
    """Arma la lista de tareas con su historial de ejecuciones ya agrupado."""
    ejecuciones_por_plan = {}
    for e in ejecuciones:
        ejecuciones_por_plan.setdefault(e.get("plan_id"), []).append(e)

    tareas = []
    for p in planes:
        if maquinas_filtradas_ids is not None and p.get("maquina_id") not in maquinas_filtradas_ids:
            continue
        tareas.append({
            "plan_id": p.get("id"),
            "tarea": p.get("tarea"),
            "maquina_id": p.get("maquina_id"),
            "frecuencia_dias": p.get("frecuencia_dias", 30),
            "ultima_ejecucion": p.get("ultima_ejecucion"),
            "ejecuciones": ejecuciones_por_plan.get(p.get("id"), [])
        })
    return tareas


# ---------------------------------------------------------------
# VISTA PREVIA INLINE (dataframe con colores, sin descargar nada)
# ---------------------------------------------------------------
def construir_dataframe_mes(tareas, map_maquina_nombre, anio, mes):
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    filas = []
    for t in tareas:
        fila = {
            "Actividad": t["tarea"],
            "Máquina": map_maquina_nombre.get(t["maquina_id"], "?"),
            "Frec. (días)": t["frecuencia_dias"]
        }
        dias_prog = _dias_programados_en_mes(t.get("ultima_ejecucion"), t["frecuencia_dias"], anio, mes)
        dias_real = _dias_realizados_en_mes(t["ejecuciones"], anio, mes)
        for d in range(1, ultimo_dia + 1):
            if d in dias_real:
                fila[str(d)] = "✓"
            elif d in dias_prog:
                fila[str(d)] = "X"
            else:
                fila[str(d)] = ""
        filas.append(fila)
    return pd.DataFrame(filas)


def _colorear_celda(val):
    if val == "✓":
        return "background-color: #22C55E; color: white; font-weight: bold; text-align: center"
    if val == "X":
        return "background-color: #38BDF8; color: white; font-weight: bold; text-align: center"
    return "text-align: center"


def mostrar_vista_previa_mes(tareas, map_maquina_nombre, anio, mes):
    if not tareas:
        st.info("No hay tareas para mostrar en este mes.")
        return
    df = construir_dataframe_mes(tareas, map_maquina_nombre, anio, mes)
    columnas_dias = [c for c in df.columns if c.isdigit()]
    st.dataframe(
        df.style.map(_colorear_celda, subset=columnas_dias),
        use_container_width=True,
        hide_index=True
    )
    st.caption("🔵 X = programado (proyección según frecuencia)   🟢 ✓ = realizado de verdad (según el historial)")


# ---------------------------------------------------------------
# GENERACIÓN DE EXCEL (una hoja por mes, o las 12 juntas)
# ---------------------------------------------------------------
def _escribir_hoja_mes(ws, tareas, map_maquina_nombre, anio, mes, nombre_empresa, logo_bytes, titulo_maquina):
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    col_actividad, col_maquina, col_frecuencia, primera_col_dia = 1, 2, 3, 4
    ultima_col_dia = primera_col_dia + ultimo_dia - 1

    fill_header = PatternFill("solid", fgColor="12171B")
    fill_finde = PatternFill("solid", fgColor="FDE68A")
    fill_abr = PatternFill("solid", fgColor="E5E9EC")
    fill_programado = PatternFill("solid", fgColor="38BDF8")
    fill_realizado = PatternFill("solid", fgColor="22C55E")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_col_dia)
    celda_titulo = ws.cell(row=1, column=1, value=f"Programa de Mantenimientos Preventivos — {titulo_maquina}")
    celda_titulo.font = Font(size=15, bold=True)
    celda_titulo.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_col_dia)
    celda_sub = ws.cell(row=2, column=1, value=f"{nombre_empresa}  ·  Mes: {MESES[mes - 1]} {anio}" if nombre_empresa else f"Mes: {MESES[mes - 1]} {anio}")
    celda_sub.font = Font(size=11, italic=True)
    celda_sub.alignment = Alignment(horizontal="center")

    if logo_bytes:
        try:
            img_buf = BytesIO(logo_bytes)
            xl_img = XLImage(img_buf)
            xl_img.width = 55
            xl_img.height = 55
            ws.add_image(xl_img, "A1")
        except Exception:
            pass

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    fila_header = 4
    ws.cell(row=fila_header, column=col_actividad, value="Actividad").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=fila_header, column=col_maquina, value="Máquina").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=fila_header, column=col_frecuencia, value="Frecuencia (días)").font = Font(bold=True, color="FFFFFF")
    for c in range(1, col_frecuencia + 1):
        ws.cell(row=fila_header, column=c).fill = fill_header
        ws.merge_cells(start_row=fila_header, start_column=c, end_row=fila_header + 1, end_column=c)
        ws.cell(row=fila_header, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    for i in range(ultimo_dia):
        dia_num = i + 1
        fecha = date(anio, mes, dia_num)
        col = primera_col_dia + i
        es_finde = fecha.weekday() >= 5

        celda_num = ws.cell(row=fila_header, column=col, value=dia_num)
        celda_num.alignment = Alignment(horizontal="center")
        celda_num.fill = fill_finde if es_finde else fill_header
        celda_num.font = Font(bold=True, size=9, color="000000" if es_finde else "FFFFFF")

        celda_abr = ws.cell(row=fila_header + 1, column=col, value=DIAS_SEMANA_ABR[fecha.weekday()])
        celda_abr.alignment = Alignment(horizontal="center")
        celda_abr.font = Font(size=8, italic=True)
        celda_abr.fill = fill_finde if es_finde else fill_abr
        ws.column_dimensions[get_column_letter(col)].width = 4

    fila_actual = fila_header + 2
    for t in tareas:
        ws.cell(row=fila_actual, column=col_actividad, value=t["tarea"])
        ws.cell(row=fila_actual, column=col_maquina, value=map_maquina_nombre.get(t["maquina_id"], "?"))
        ws.cell(row=fila_actual, column=col_frecuencia, value=t["frecuencia_dias"])

        dias_prog = _dias_programados_en_mes(t.get("ultima_ejecucion"), t["frecuencia_dias"], anio, mes)
        dias_real = _dias_realizados_en_mes(t["ejecuciones"], anio, mes)
        for i in range(ultimo_dia):
            dia_num = i + 1
            col = primera_col_dia + i
            if dia_num in dias_real:
                celda = ws.cell(row=fila_actual, column=col, value="✓")
                celda.fill = fill_realizado
                celda.alignment = Alignment(horizontal="center")
                celda.font = Font(bold=True, color="FFFFFF")
            elif dia_num in dias_prog:
                celda = ws.cell(row=fila_actual, column=col, value="X")
                celda.fill = fill_programado
                celda.alignment = Alignment(horizontal="center")
                celda.font = Font(bold=True, color="FFFFFF")
        fila_actual += 1

    fila_leyenda = fila_actual + 1
    ws.cell(row=fila_leyenda, column=1, value="🔵 X = programado (proyección)   🟢 ✓ = realizado según historial").font = Font(size=9, italic=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = ws.cell(row=fila_header + 2, column=primera_col_dia)


def generar_excel_calendario_mes(tareas, map_maquina_nombre, anio, mes, nombre_empresa, logo_bytes, titulo_maquina):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = MESES[mes - 1][:28]
    _escribir_hoja_mes(ws, tareas, map_maquina_nombre, anio, mes, nombre_empresa, logo_bytes, titulo_maquina)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_calendario_anual(tareas, map_maquina_nombre, anio, nombre_empresa, logo_bytes, titulo_maquina):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for mes in range(1, 13):
        ws = wb.create_sheet(title=MESES[mes - 1][:28])
        _escribir_hoja_mes(ws, tareas, map_maquina_nombre, anio, mes, nombre_empresa, logo_bytes, titulo_maquina)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------
# UI
# ---------------------------------------------------------------
def render_calendario():
    st.title("📅 Calendario de Mantenimiento Preventivo")
    st.write("Vista previa acá mismo, y descarga en Excel si la necesitás para imprimir o compartir.")

    maquinas = st.session_state.get("maquinas", [])
    planes = st.session_state.get("planes", [])
    ejecuciones = st.session_state.get("plan_ejecuciones", [])

    if not maquinas:
        st.warning("Registrá al menos una máquina primero.")
        return
    if not planes:
        st.warning("Todavía no hay tareas cargadas en '🗓️ Plan Preventivo'. Cargá al menos una para poder generar el calendario.")
        return

    map_maquina_nombre = {m["id"]: m["nombre"] for m in maquinas}
    dict_maquinas = {m["nombre"]: m["id"] for m in maquinas}
    opciones_maquina = ["Todas las Máquinas"] + list(dict_maquinas.keys())

    col1, col2, col3 = st.columns(3)
    maquina_sel = col1.selectbox("Máquina", opciones_maquina)
    anio_sel = col2.number_input("Año", min_value=2020, max_value=2100, value=datetime.now().year, step=1)
    ver_anio_completo = col3.checkbox("Ver año completo", value=False)

    if not ver_anio_completo:
        mes_sel = st.selectbox("Mes", list(range(1, 13)), index=datetime.now().month - 1, format_func=lambda m: MESES[m - 1])

    config_empresa = st.session_state.get("config_empresa", {}) or {}
    nombre_empresa = config_empresa.get("nombre_empresa", "") or ""
    logo_bytes = base64.b64decode(config_empresa["logo_base64"]) if config_empresa.get("logo_base64") else None
    if not nombre_empresa and not logo_bytes:
        st.caption("💡 Podés cargar el logo y nombre de tu empresa una sola vez en '📑 Reportes' → Membrete.")

    maquinas_filtradas_ids = None if maquina_sel == "Todas las Máquinas" else {dict_maquinas[maquina_sel]}
    titulo_maquina = "Todas las Máquinas" if maquina_sel == "Todas las Máquinas" else maquina_sel
    tareas = _preparar_tareas(planes, ejecuciones, maquinas_filtradas_ids)

    if not tareas:
        st.warning("No hay tareas de plan preventivo para esa máquina.")
        return

    st.markdown("---")
    st.markdown("##### 👁️ Vista previa")

    if ver_anio_completo:
        tabs = st.tabs(MESES)
        for i, tab in enumerate(tabs):
            with tab:
                mostrar_vista_previa_mes(tareas, map_maquina_nombre, anio_sel, i + 1)

        st.markdown("---")
        if st.button("📥 Generar Excel del Año Completo (12 hojas)"):
            buffer = generar_excel_calendario_anual(tareas, map_maquina_nombre, anio_sel, nombre_empresa, logo_bytes, titulo_maquina)
            st.download_button(
                "⬇️ Descargar Calendario_Anual.xlsx",
                data=buffer,
                file_name=f"Calendario_Preventivo_Anual_{anio_sel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ Calendario anual generado (una hoja por mes).")
    else:
        mostrar_vista_previa_mes(tareas, map_maquina_nombre, anio_sel, mes_sel)

        st.markdown("---")
        if st.button("📥 Generar Excel de este Mes"):
            buffer = generar_excel_calendario_mes(tareas, map_maquina_nombre, anio_sel, mes_sel, nombre_empresa, logo_bytes, titulo_maquina)
            st.download_button(
                "⬇️ Descargar Calendario_Preventivo.xlsx",
                data=buffer,
                file_name=f"Calendario_Preventivo_{anio_sel}_{mes_sel:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ Calendario generado.")
