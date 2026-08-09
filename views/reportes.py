import streamlit as st
import pandas as pd
import base64
from io import BytesIO
from datetime import datetime
from views.dashboard import calcular_kpis_industriales
from database.conection import guardar_configuracion_empresa


def generar_pdf_maquinas(maquinas, nombre_empresa="", logo_bytes=None):
    """Genera un PDF con membrete (logo + nombre de empresa) y el listado de
    máquinas ordenado por criticidad, para el informe de gerencia."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = []

    if logo_bytes:
        try:
            logo_buf = BytesIO(logo_bytes)
            img = Image(logo_buf, width=3 * cm, height=3 * cm)
            img.hAlign = "LEFT"
            story.append(img)
            story.append(Spacer(1, 8))
        except Exception:
            pass

    titulo_style = ParagraphStyle("TituloEmpresa", parent=styles["Title"], fontSize=16, spaceAfter=2)
    subtitulo_style = ParagraphStyle("Subtitulo", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    if nombre_empresa:
        story.append(Paragraph(nombre_empresa, titulo_style))
    story.append(Paragraph("Listado de Máquinas y Criticidad", styles["Heading2"]))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitulo_style))
    story.append(Spacer(1, 16))

    orden_crit = {"A": 0, "B": 1, "C": 2}
    maquinas_ordenadas = sorted(maquinas, key=lambda m: orden_crit.get(m.get("criticidad"), 3))

    crit_texto = {"A": "Crítica (A)", "B": "Importante (B)", "C": "Menor (C)"}
    data = [["Máquina", "Código", "Sección", "Criticidad"]]
    for m in maquinas_ordenadas:
        data.append([
            m.get("nombre", "—"),
            m.get("codigo", "—"),
            m.get("seccion") or "—",
            crit_texto.get(m.get("criticidad"), m.get("criticidad", "—"))
        ])

    tabla = Table(data, colWidths=[5.5 * cm, 3 * cm, 4 * cm, 3.5 * cm], repeatRows=1)
    estilo = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12171B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])
    for i, m in enumerate(maquinas_ordenadas, start=1):
        if m.get("criticidad") == "A":
            estilo.add("TEXTCOLOR", (3, i), (3, i), colors.HexColor("#B91C1C"))
            estilo.add("FONTNAME", (3, i), (3, i), "Helvetica-Bold")
    tabla.setStyle(estilo)
    story.append(tabla)

    story.append(Spacer(1, 20))
    resumen_txt = (
        f"Total de máquinas: {len(maquinas)}  ·  "
        f"Críticas (A): {len([m for m in maquinas if m.get('criticidad') == 'A'])}  ·  "
        f"Importantes (B): {len([m for m in maquinas if m.get('criticidad') == 'B'])}  ·  "
        f"Menores (C): {len([m for m in maquinas if m.get('criticidad') == 'C'])}"
    )
    story.append(Paragraph(resumen_txt, styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _construir_costos_por_maquina(ots, maquinas, ot_repuestos, repuestos):
    """Costo total (mano de obra + repuestos consumidos) agrupado por máquina."""
    map_maquina = {m["id"]: m["nombre"] for m in maquinas}
    map_repuesto_costo = {r["id"]: r.get("costo_unitario", 0) for r in repuestos}

    costo_repuestos_por_ot = {}
    for orr in ot_repuestos:
        ot_id = orr.get("ot_id")
        costo = orr.get("cantidad_usada", 0) * map_repuesto_costo.get(orr.get("repuesto_id"), 0)
        costo_repuestos_por_ot[ot_id] = costo_repuestos_por_ot.get(ot_id, 0) + costo

    filas = {}
    for o in ots:
        m_nombre = map_maquina.get(o.get("maquina_id"), "Desconocida")
        costo_mo = o.get("costo_mano_obra", 0) or 0
        costo_rep = costo_repuestos_por_ot.get(o.get("id"), 0)
        horas_paro = o.get("horas_paro", 0) or 0

        if m_nombre not in filas:
            filas[m_nombre] = {"Máquina": m_nombre, "OTs": 0, "Costo Mano de Obra": 0, "Costo Repuestos": 0, "Costo Total": 0, "Horas de Paro": 0}

        filas[m_nombre]["OTs"] += 1
        filas[m_nombre]["Costo Mano de Obra"] += costo_mo
        filas[m_nombre]["Costo Repuestos"] += costo_rep
        filas[m_nombre]["Costo Total"] += costo_mo + costo_rep
        filas[m_nombre]["Horas de Paro"] += horas_paro

    return sorted(filas.values(), key=lambda x: x["Costo Total"], reverse=True)


def construir_hojas_reporte(maquinas, fallas, ots, repuestos, terceros, planes, kpis):
    """Arma los dataframes de cada hoja, reusados tanto en la vista previa
    en pantalla como al generar el Excel final."""
    map_maquina = {m["id"]: m["nombre"] for m in maquinas}

    resumen = pd.DataFrame([{
        "Fecha del Reporte": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Máquinas Registradas": len(maquinas),
        "Disponibilidad (%)": kpis["disponibilidad"],
        "MTTR (hrs)": kpis["mttr"],
        "MTBF (hrs)": kpis["mtbf"],
        "Horas de Paro Totales": kpis["horas_paro"],
        "Fallas Abiertas": len([f for f in fallas if f.get("estado") != "Cerrada"]),
        "Repuestos Críticos": len([r for r in repuestos if r.get("stock_actual", 0) <= r.get("stock_minimo", 0)])
    }])

    backlog_rows = [{
        "Código": o.get("codigo"),
        "Máquina": map_maquina.get(o.get("maquina_id"), "Desconocida"),
        "Tipo": o.get("tipo_mantenimiento"),
        "Estado": o.get("estado"),
        "Descripción": o.get("descripcion"),
        "Horas de Paro": o.get("horas_paro")
    } for o in ots if o.get("estado") != "Completada"]
    backlog = pd.DataFrame(backlog_rows if backlog_rows else [{"Info": "Sin pendientes"}])

    costos_rows = _construir_costos_por_maquina(ots, maquinas, st.session_state.get("ot_repuestos", []), repuestos)
    costos = pd.DataFrame(costos_rows if costos_rows else [{"Info": "Sin datos de costos aún"}])

    criticos_rows = [{
        "Repuesto": r.get("nombre"),
        "Código Interno": r.get("codigo_interno"),
        "Stock Actual": r.get("stock_actual"),
        "Stock Mínimo": r.get("stock_minimo")
    } for r in repuestos if r.get("stock_actual", 0) <= r.get("stock_minimo", 0)]
    criticos = pd.DataFrame(criticos_rows if criticos_rows else [{"Info": "Sin repuestos críticos"}])

    inventario_rows = [{
        "Repuesto": r.get("nombre"),
        "Código Interno": r.get("codigo_interno"),
        "Stock Actual": r.get("stock_actual", 0),
        "Stock Mínimo": r.get("stock_minimo", 0),
        "Costo Unitario (Gs.)": r.get("costo_unitario", 0),
        "Valor Total (Gs.)": (r.get("stock_actual", 0) or 0) * (r.get("costo_unitario", 0) or 0)
    } for r in repuestos]
    inventario = pd.DataFrame(inventario_rows if inventario_rows else [{"Info": "Sin repuestos cargados"}])

    hoy = datetime.now().date()
    actividades_por_plan = {}
    for a in st.session_state.get("plan_actividades", []):
        actividades_por_plan.setdefault(a.get("plan_id"), []).append(a.get("actividad"))

    plan_rows = []
    for p in planes:
        try:
            fecha_prox = datetime.strptime(p.get("proxima_ejecucion"), "%Y-%m-%d").date()
            dias = (fecha_prox - hoy).days
        except (TypeError, ValueError):
            dias = None
        plan_rows.append({
            "Máquina": map_maquina.get(p.get("maquina_id"), "Desconocida"),
            "Plan": p.get("nombre_plan"),
            "Actividades": "; ".join(actividades_por_plan.get(p.get("id"), [])) or "—",
            "Frecuencia (días)": p.get("frecuencia_dias"),
            "Próxima Ejecución": p.get("proxima_ejecucion"),
            "Días Restantes": dias
        })
    plan_df = pd.DataFrame(plan_rows if plan_rows else [{"Info": "Sin plan preventivo cargado"}])

    venc_rows = [{
        "Equipo": t.get("nombre"),
        "Proveedor": t.get("contacto"),
        "Próximo Vencimiento": t.get("proximoVencimiento")
    } for t in terceros]
    terceros_df = pd.DataFrame(venc_rows if venc_rows else [{"Info": "Sin terceros cargados"}])

    return {
        "Resumen Ejecutivo": resumen,
        "Backlog OTs": backlog,
        "Costos por Máquina": costos,
        "Inventario Completo": inventario,
        "Repuestos Críticos": criticos,
        "Plan Preventivo": plan_df,
        "Terceros": terceros_df
    }


def generar_excel_reporte(hojas: dict, nombre_empresa="", logo_bytes=None):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            fila_inicio = 5 if (nombre_hoja == "Resumen Ejecutivo" and (nombre_empresa or logo_bytes)) else 0
            df.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=fila_inicio)

            if nombre_hoja == "Resumen Ejecutivo" and (nombre_empresa or logo_bytes):
                ws = writer.sheets[nombre_hoja]
                if nombre_empresa:
                    ws["C1"] = nombre_empresa
                    ws["C1"].font = Font(size=16, bold=True)
                ws["C2"] = "Reporte de Mantenimiento"
                ws["C2"].font = Font(size=12, bold=True)
                ws["C3"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                if logo_bytes:
                    img_buf = BytesIO(logo_bytes)
                    xl_img = XLImage(img_buf)
                    xl_img.width = 70
                    xl_img.height = 70
                    ws.add_image(xl_img, "A1")
    buffer.seek(0)
    return buffer


def render_reportes():
    st.title("📑 Reportes para Gerencia")
    st.write("Mirá los datos acá mismo — descargalos solo si necesitás el archivo.")

    maquinas = st.session_state.get("maquinas", [])
    fallas = st.session_state.get("fallas", [])
    ots = st.session_state.get("ots", [])
    repuestos = st.session_state.get("repuestos", [])
    terceros = st.session_state.get("terceros", [])
    planes = st.session_state.get("planes", [])

    kpis = calcular_kpis_industriales(maquinas, fallas, ots)

    st.markdown("##### 🖼️ Membrete de la empresa")
    st.caption("Se guarda una sola vez en la base de datos — no hace falta volver a subir el logo cada vez.")

    config_actual = st.session_state.get("config_empresa", {}) or {}
    nombre_empresa = st.text_input(
        "Nombre de la empresa",
        value=config_actual.get("nombre_empresa", "") or ""
    )
    logo_file = st.file_uploader("Logo de la empresa (PNG o JPG) — subilo solo si querés cambiarlo", type=["png", "jpg", "jpeg"])

    if st.button("💾 Guardar Membrete"):
        if logo_file:
            logo_base64_nuevo = base64.b64encode(logo_file.read()).decode("utf-8")
        else:
            logo_base64_nuevo = config_actual.get("logo_base64", "")
        guardar_configuracion_empresa(nombre_empresa, logo_base64_nuevo)
        st.session_state.config_empresa = {"nombre_empresa": nombre_empresa, "logo_base64": logo_base64_nuevo}
        st.success("✅ Membrete guardado. Ya no hace falta volver a cargarlo.")
        st.rerun()

    logo_bytes = base64.b64decode(config_actual["logo_base64"]) if config_actual.get("logo_base64") else None

    # --- VISTA PREVIA SIEMPRE VISIBLE, EN PESTAÑAS (sin necesidad de descargar) ---
    st.markdown("---")
    st.markdown("##### 👁️ Vista previa del reporte")

    hojas = construir_hojas_reporte(maquinas, fallas, ots, repuestos, terceros, planes, kpis)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Disponibilidad", f"{kpis['disponibilidad']}%")
    c2.metric("MTTR", f"{kpis['mttr']} hrs")
    c3.metric("MTBF", f"{kpis['mtbf']} hrs")
    c4.metric("Horas de Paro Totales", f"{kpis['horas_paro']} hrs")

    tabs = st.tabs(list(hojas.keys()))
    for tab, (nombre_hoja, df) in zip(tabs, hojas.items()):
        with tab:
            st.dataframe(df, use_container_width=True, hide_index=True)

    # --- GRÁFICOS ---
    st.markdown("---")
    st.markdown("##### 📊 Gráficos")

    col_g1, col_g2 = st.columns(2)

    with col_g1:
        st.markdown("**Costo total por máquina**")
        df_costos = hojas["Costos por Máquina"]
        if "Costo Total" in df_costos.columns and not df_costos.empty:
            st.bar_chart(df_costos.set_index("Máquina")["Costo Total"])
        else:
            st.caption("Todavía no hay costos cargados en las OTs.")

    with col_g2:
        st.markdown("**OTs por estado**")
        if ots:
            conteo_estado = pd.Series([o.get("estado", "Sin estado") for o in ots]).value_counts()
            st.bar_chart(conteo_estado)
        else:
            st.caption("Todavía no hay OTs cargadas.")

    col_g3, col_g4 = st.columns(2)

    with col_g3:
        st.markdown("**Fallas por estado**")
        if fallas:
            conteo_fallas = pd.Series([f.get("estado", "Sin estado") for f in fallas]).value_counts()
            st.bar_chart(conteo_fallas)
        else:
            st.caption("Todavía no hay fallas registradas.")

    with col_g4:
        st.markdown("**Stock actual vs. mínimo (repuestos críticos)**")
        df_criticos_chart = hojas["Repuestos Críticos"]
        if "Repuesto" in df_criticos_chart.columns and not df_criticos_chart.empty:
            st.bar_chart(df_criticos_chart.set_index("Repuesto")[["Stock Actual", "Stock Mínimo"]])
        else:
            st.caption("No hay repuestos en nivel crítico ahora mismo. 🎉")

    if st.button("📥 Generar y Descargar Excel"):
        buffer = generar_excel_reporte(hojas, nombre_empresa, logo_bytes)
        st.download_button(
            label="⬇️ Descargar Reporte_Mantenimiento.xlsx",
            data=buffer,
            file_name=f"Reporte_Mantenimiento_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("✅ Excel listo para descargar.")

    # --- LISTADO DE MÁQUINAS EN PDF: previsualizado inline, sin forzar descarga ---
    st.markdown("---")
    st.markdown("##### 📄 Listado de Máquinas y Criticidad (PDF con membrete)")
    st.caption("Ideal para el primer informe a gerencia — usa el mismo nombre y logo que cargaste arriba.")

    if st.button("👁️ Generar Vista Previa del PDF"):
        pdf_buffer = generar_pdf_maquinas(maquinas, nombre_empresa, logo_bytes)
        st.session_state["_pdf_preview_bytes"] = pdf_buffer.getvalue()

    if st.session_state.get("_pdf_preview_bytes"):
        import pymupdf  # import perezoso: solo pesa en memoria si de verdad se usa esta función
        # Chrome bloquea los iframes con PDF embebido como data:URI cuando la
        # página ya está dentro de otro iframe (como pasa en Streamlit Cloud).
        # Por eso mostramos cada página del PDF como imagen, no como iframe.
        doc_pdf = pymupdf.open(stream=st.session_state["_pdf_preview_bytes"], filetype="pdf")
        for i, pagina in enumerate(doc_pdf):
            pix = pagina.get_pixmap(dpi=150)
            st.image(pix.tobytes("png"), use_container_width=True, caption=f"Página {i + 1} de {len(doc_pdf)}")
        doc_pdf.close()

        st.download_button(
            label="⬇️ Descargar Listado_Maquinas.pdf",
            data=st.session_state["_pdf_preview_bytes"],
            file_name=f"Listado_Maquinas_{datetime.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf"
        )
